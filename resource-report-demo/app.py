#!/usr/bin/env python3
# Author: Ryan Tiffany
# Copyright (c) 2024
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#     http://www.apache.org/licenses/LICENSE-2.0
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

import os
import json
import click
import logging
from ibm_platform_services import IamIdentityV1, ResourceControllerV2, ResourceManagerV2
from ibm_platform_services.resource_controller_v2 import ResourceInstancesPager, ResourceBindingsPager, ResourceKeysPager
from ibm_cloud_sdk_core.authenticators import IAMAuthenticator
from ibm_cloud_sdk_core import ApiException
from datetime import datetime
import pandas as pd
from icecream import ic
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

ibmcloud_api_key = os.environ.get('IBMCLOUD_API_KEY')
if not ibmcloud_api_key:
    raise ValueError("IBMCLOUD_API_KEY environment variable not found")

authenticator = IAMAuthenticator(apikey=ibmcloud_api_key)

def setup_logging(default_path='logging.json', default_level=logging.info, env_key='LOG_CFG'):
    path = default_path
    value = os.getenv(env_key, None)
    if value:
        path = value
    if os.path.exists(path):
        with open(path, 'rt') as f:
            config = json.load(f)
        logging.config.dictConfig(config)
    else:
        logging.basicConfig(level=default_level)

def iam_client():  
    iamIdentityService = IamIdentityV1(authenticator=authenticator)
    return iamIdentityService

def resource_controller_client():  
    resourceControllerService = ResourceControllerV2(authenticator=authenticator)
    return resourceControllerService

def resource_manager_client():  
    resourceManagerService = ResourceManagerV2(authenticator=authenticator)
    return resourceManagerService



def get_account_id():
    try:
        client = iam_client()
        api_key = client.get_api_keys_details(
          iam_api_key=ibmcloud_api_key
        ).get_result()
    except ApiException as e:
        logging.error("API exception {}.".format(str(e)))
        quit(1)
    account_id = api_key["account_id"]
    return account_id

def extract_service_name(crn):
    # The pattern looks for 'public:' followed by any characters until the next ':' or the end of the string
    pattern = r"public:([^:]*)"
    match = re.search(pattern, crn)
    if match:
        # The first group is the part we're interested in
        return match.group(1)
    else:
        return None

def extract_is_type(crn):
    # The pattern looks for '::' followed by any characters until the next ':' or the end of the string
    pattern = r"::([^:]*)"
    match = re.search(pattern, crn)
    if match:
        # The first group is the part we're interested in
        return match.group(1)
    else:
        return None



@click.group()
def cli():
    """Group to hold our commands"""
    pass

@cli.command()
def get_account_resources():
    client = resource_controller_client()
    all_results = []
    pager = ResourceInstancesPager(
        client=client
    )
    while pager.has_next():
        next_page = pager.get_next()
        assert next_page is not None
        all_results.extend(next_page)

    results = ic(all_results)
    return results


@cli.command()
def get_resource_service_keys():
    client = resource_controller_client()
    all_results = []
    pager = ResourceKeysPager(
        client=client
    )
    while pager.has_next():
        next_page = pager.get_next()
        assert next_page is not None
        all_results.extend(next_page)

    results = ic(all_results)
    return results

## need to add try statements here with explicit error handling
@cli.command()
def get_resource_groups():
    account_id = get_account_id()
    client = resource_manager_client()
    try:
        resource_group_list = client.list_resource_groups(
            account_id=account_id
        ).get_result()
        return resource_group_list
    
    except ApiException as e:
        logging.error("API exception {}.".format(str(e)))
        quit(1)


@cli.command()
def get_resources_by_group():
    rc_client = resource_controller_client()
    resource_groups = get_resource_groups()
    
    # Create a new Excel writer object
    writer = pd.ExcelWriter('resource_groups.xlsx', engine='openpyxl')
    
    for group in resource_groups.get("resources"):
        resource_group_id = group["id"]
        pager = ResourceInstancesPager(
            client=rc_client,
            resource_group_id=resource_group_id
            )
        all_results = []
        while pager.has_next():
            next_page = pager.get_next()
            assert next_page is not None
            all_results.extend(next_page)
        
        resource_instances = []
        for result in all_results:
            crn_slug = extract_service_name(result.get('crn')) if 'crn' in result else None
            is_type = extract_is_type(result.get('crn')) if 'crn' in result and crn_slug == 'is' else None

            resource_instance = {
                'name': result.get('name'),
                'type': result.get('type'),
                'resource_id': result.get('resource_id'),
                'crn_slug': crn_slug,
                'created_by': result.get('created_by'),
                'crn': result.get('crn')
            }

            if is_type:
                resource_instance['is_type'] = is_type
            
            resource_instances.append(resource_instance)
        
        df = pd.DataFrame(resource_instances)
        
        df.to_excel(writer, sheet_name=group["name"], index=False)
    
    writer.save()


@cli.command()
def write_resources_by_group():
    rc_client = resource_controller_client()
    resource_groups = get_resource_groups()
    
    # Create a new Excel writer object
    writer = pd.ExcelWriter('resource_groups.xlsx', engine='openpyxl')
    
    for group in resource_groups.get("resources"):
        resource_group_id = group["id"]
        pager = ResourceInstancesPager(
            client=rc_client,
            resource_group_id=resource_group_id
            )
        all_results = []
        while pager.has_next():
            next_page = pager.get_next()
            assert next_page is not None
            all_results.extend(next_page)
        
        resource_instances = []
        for result in all_results:
            crn_slug = extract_service_name(result.get('crn')) if 'crn' in result else None
            is_type = extract_is_type(result.get('crn')) if 'crn' in result and crn_slug == 'is' else None

            resource_instance = {
                'name': result.get('name'),
                'type': result.get('type'),
                'resource_id': result.get('resource_id'),
                'crn_slug': crn_slug,
                'created_by': result.get('created_by'),
                'crn': result.get('crn')
            }

            if is_type:
                resource_instance['is_type'] = is_type
            
            resource_instances.append(resource_instance)
        
        # Convert the list of dictionaries to a DataFrame
        df = pd.DataFrame(resource_instances)
        
        # Write the DataFrame to an Excel sheet named after the resource group
        df.to_excel(writer, sheet_name=group["name"], index=False)
    
    # Save the Excel file
    writer.save()


@cli.command()
def write_excel():
    account_id = get_account_id()
    rc_client = resource_controller_client()
    rm_client = resource_manager_client()
    resource_group_list = rm_client.list_resource_groups(
        account_id=account_id
    ).get_result()
    
    # Create a new workbook
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    
    for group in resource_group_list["resources"]:
        resource_group_id = group["id"]
        resource_group_name = group["name"]
        resource_list = rc_client.list_resource_instances(
            resource_group_id=resource_group_id
        ).get_result()
        
        result = resource_list.get("resources")
        crn_slug = extract_service_name(result.get('crn')) if 'crn' in result else None
        is_type = extract_is_type(result.get('crn')) if 'crn' in result and crn_slug == 'is' else None
        if is_type:
            result['is_type'] = is_type
            # Create a new sheet for each resource group
        ws = wb.create_sheet(title=resource_group_name)
        
        # Write the headers
        headers = ["Resource Name", "Resource Type", "Resource ID", "CRN Slug", "Created By", "IS Type"]
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = column_title
        
        # Write the resources
        for row_num, resource in enumerate(result, 2):
            ws.cell(row=row_num, column=1, value=resource["name"])
            ws.cell(row=row_num, column=2, value=resource["resource_id"])
            ws.cell(row=row_num, column=3, value=resource["id"])
            ws.cell(row=row_num, column=4, value=crn_slug)
            ws.cell(row=row_num, column=5, value=resource["created_by"])
            ws.cell(row=row_num, column=6, value=is_type)
    
    # Save the workbook
    wb.save("resource_groups.xlsx")



if __name__ == '__main__':
    cli()